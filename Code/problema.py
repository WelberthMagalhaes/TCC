import pandas as pd
import random
import openpyxl


class Problema:

	def __init__(self, arq):

		def geraAleatorios(caminho):
			book = openpyxl.load_workbook(caminho,data_only=True)
			sheet1 = book['home']
			T = int(sheet1['B5'].value)			
			
			#N - número de toners
			N=[]
			i = 2
			aux = sheet1.cell(row=1,column=i).value
			
			while(aux != None):
				N.append(aux)
				i+=1
				aux = sheet1.cell(row=1,column=i).value

			#estoque inicial
			for i in range(2, len(N)+2):
				sheet1.cell(row=3,column=i).value = random.randint(1,20)				
				aux = sheet1.cell(row=3,column=i).value * (random.randint(1,5))
				sheet1.cell(row=3,column=i).value = aux
			#------
			
			sheet4 = book['demandaToner']
			sheet5 = book['custoCapital']
			sheet6 = book['custoToner']

			indice = 1
			
			for i in range(2,T+2):
				#demanda toner
				sheet4.cell(row=1,column=i).value = "t%d"%(indice)				
				#custo toner
				sheet6.cell(row=1,column=i).value = "t%d"%(indice)
				#custo capital
				sheet5.cell(row=1,column=i).value = "t%d"%(indice)

				for k in range(2, len(N)+2):
					#custo capital
					#=custoToner!C9*((1+home!$B$10)^(1/12)-1)
					aux = sheet6.cell(row=k,column=i).value * ((1+sheet1.cell(row=10,column=2).value)**(1/12)-1)
					sheet5.cell(row=k,column=i).value = aux
				
				
				for j in range(2,len(N)+2):
					if(i>T):break
					
					#demanda toner
					aux = sheet4.cell(row=j,column=2).value * random.randint(1,T)#ok
					sheet4.cell(row=j,column=i+1).value = aux
					
					#custo toner
					aux = sheet6.cell(row=j,column=2).value * random.uniform(0.6,1.4)#ok
					sheet6.cell(row=j,column=i+1).value = aux

				indice += 1

			aleatorio='C:/Users/wheidermagal/OneDrive - DXC Production/Documents/Pessoal/TCC/inputs/aleatorio.xlsx'
			book.save(aleatorio)
			book.close
			return aleatorio

		#gerando dados aleatórios
		#arq = geraAleatorios(arq)
		
		#leitura do arquivo
		df = pd.read_excel(arq, sheet_name=0)
		book = openpyxl.load_workbook(arq, data_only=True)

		# N - Conjunto de Toners
		aux = df.columns.to_list()
		aux.pop(0)
		self.N = aux
		#--------------------

		# T - Número de períodos
		self.T = int(df.iat[3,1])
		#--------------------

		# A - Capacidade de Armazenamento
		self.A = df.iat[5,1]
		#--------------------

		#si - Estoque inicial do toner i
		s = df
		s = s.T
		s = s.drop('Conjunto de Toners')
		s = s.drop(columns=0)
		ind = s.index.to_list()
		self.si = {ind[i] : s.values[i][0] for i in range(s[1].size)}
		#--------------------
		
		col = int(self.T+1)
		# ci - Custo unitário do toner i
		df = pd.read_excel(arq, sheet_name=5)
		self.ci = {df.values[i][0] : df.values[i][1:col] for i in range(len(self.N))}
		#--------------------

		# vi - Volume ocupado por cada toner #Validado 16/12
		df = pd.read_excel(arq, sheet_name=1)
		self.vi = {df.values[i][0] : df.values[i][1] for i in range(len(self.N))}
		#--------------------

		# ei - margem mínima média do toner i #Validado 16/12
		df = pd.read_excel(book, sheet_name=2,engine="openpyxl")
		self.ei = {df.values[i][0] : df.values[i][1] for i in range(len(self.N))}
		#--------------------

		
		# dit - demanda média esperada do toner i no período t  #Validado 16/12
		df = pd.read_excel(arq, sheet_name=3)
		self.di = {df.values[i][0] : df.values[i][1:col] for i in range(len(self.N))}
		#--------------------

		# hi - Custo de capital do estoque por unidade do toner i
		df = pd.read_excel(book, sheet_name=4,engine="openpyxl")
		self.hi = {df.values[i][0] : df.values[i][1:col] for i in range(len(self.N))}
		#--------------------